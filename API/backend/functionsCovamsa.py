from flask import Flask, jsonify, request, Response

import os
from os import path
import base64
import pymysql.cursors

import base64

import json
import sys
import copy

import time
import datetime
import openpyxl

strMysqluUser = "root"
strMysqlHost="localhost"
strMysqlPort=3306
strMysqlPassword="cvm980423k63"
strMysqlDB="covamsa_prov"

def fnGetServicioPesado():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		cursor.callproc('getServicioPesado')
		response = cursor.fetchall()
		result=[]
		for row in response:
			aux=""
			if row['PosicionDel'] == 1 and row['PosicionTra'] == 1 and row['PosicionCab'] == 1:
				aux="Del Tra Cab"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 1 and row['PosicionCab'] == 0:
				aux="Del Tra"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 0 and row['PosicionCab'] == 1:
				aux="Del Cab"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 0 and row['PosicionCab'] == 0:
				aux="Del"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 1 and row['PosicionCab'] == 1:
				aux="Tra Cab"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 1 and row['PosicionCab'] == 0:
				aux="Tra"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 0 and row['PosicionCab'] == 1:
				aux="Cab"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 0 and row['PosicionCab'] == 0:
				aux="Ninguna"

			if(row['TipoMontajeInf'] == "A"):
				auxInf="Anillo"
			elif(row['TipoMontajeInf'] == "AB"):
				auxInf="Anillo - Buje"
			elif(row['TipoMontajeInf'] == "ABC"):
				auxInf="Anillo - Buje - Camisa"
			elif(row['TipoMontajeInf'] == "P"):
				auxInf="Perno"
			elif(row['TipoMontajeInf'] == "ATB"):
				auxInf="Anillo - Buje Trapezoidal"
			
			if(row['TipoMontajeSup'] == "A"):
				auxSup="Anillo"
			elif(row['TipoMontajeSup'] == "AB"):
				auxSup="Anillo - Buje"
			elif(row['TipoMontajeSup'] == "ABC"):
				auxSup="Anillo - Buje - Camisa"
			elif(row['TipoMontajeSup'] == "P"):
				auxSup="Perno"
			elif(row['TipoMontajeSup'] == "ATB"):
				auxSup="Anillo - Buje Trapezoidal"
			obj={
				'ID': row['ID'],
				'GABRIEL': row['GABRIEL'],
				'MONROE': row['MONROE'], 
				'GRC': row['GRC'], 
				'Armadora': row['Armadora'], 
				'Posicion': aux, 
				'Tipo': row['Tipo'], 
				'LongitudExp': row['LongitudExp'], 
				'LongitudComp': row['LongitudComp'], 
				'Carrera': row['Carrera'], 
				'TipoMontajeSup': auxSup, 
				'DiametroSup': row['DiametroSup'], 
				'LongitudSup': row['LongitudSup'], 
				'TipoMontajeInf': auxInf, 
				'DiametroInf': row['DiametroInf'], 
				'LongitudInf': row['LongitudInf'],
				"info": row['info']
			}
			result.append(obj)

		if result:
			#print(result)
			return {'intStatus':200, 'strAnswer': result}
		else:
			return {'intStatus':200, 'strAnswer': 'No hay nada'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()



def fnPostServicioPesado(idGabriel,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			idGabriel,
			info
		)
		cursor.callproc('postServicioPesadoInfo',params)
		MysqlCnx.commit()

		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostLogin(usuario,password):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			usuario,
			password
		)
		cursor.callproc('getLogin',params)
		response = cursor.fetchall()
		#print(response)
		if response:
			#print(result)
			return {'intStatus':200, 'strAnswer': usuario}
		else:
			return {'intStatus':202, 'strAnswer': 'Credenciales incorrectas.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetBolsasDeAire():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		cursor.callproc('getBolsasDeAire')
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostBolsasDeAire(idGabriel,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			idGabriel,
			info
		)
		cursor.callproc('postBolsasDeAireInfo',params)
		MysqlCnx.commit()

		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetServicioLigero():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		cursor.callproc('getServicioLigero')
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostServicioLigero(ID,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			ID,
			info
		)
		cursor.callproc('postServicioLigeroInfo',params)
		MysqlCnx.commit()

		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetMuelles(marca):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()

		if marca == "all":
			cursor.callproc('getMuellesAll')
		else:
			params = (
				marca,
			)
			cursor.callproc('getMuelles',params)

		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostMuelles(ID,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			ID,
			info
		)
		cursor.callproc('postMuellesInfo',params)
		MysqlCnx.commit()

		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetMuellesMarcas():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		cursor.callproc('getMuellesMarcas')
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetColumns(tabla):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		sentence = "SHOW COLUMNS FROM "+tabla
		cursor.execute(sentence)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def	fnAddProductSP(GABRIEL,MONROE,GRC,Armadora,Posicion,Tipo,LongitudExp,LongitudComp,Carrera,TipoMontajeSup,DiametroSup,LongitudSup,TipoMontajeInf,DiametroInf,LongitudInf,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()

		if Posicion == "DEL":
			PosicionDel = 1
			PosicionTra = 0
			PosicionCab = 0
		elif Posicion == "TRA":
			PosicionDel = 0
			PosicionTra = 1
			PosicionCab = 0
		elif Posicion == "CAB":
			PosicionDel = 0
			PosicionTra = 0
			PosicionCab = 1
		elif Posicion == "DEL TRA":
			PosicionDel = 1
			PosicionTra = 1
			PosicionCab = 0
		
		params = (
			GABRIEL,
			MONROE,
			GRC,
			Armadora,
			PosicionDel,
			PosicionTra,
			PosicionCab,
			Tipo,
			LongitudExp,
			LongitudComp,
			Carrera,
			TipoMontajeSup,
			DiametroSup,
			LongitudSup,
			TipoMontajeInf,
			DiametroInf,
			LongitudInf,
			info
		)
		cursor.callproc('postAddProductSP',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()
		
if __name__ == '__main__':
    fnGetServicioPesado()
	
def fnGetServicioPesadoID(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('getServicioPesadoID',params)
		response = cursor.fetchall()
		result=[]
		for row in response:
			aux=""
			if row['PosicionDel'] == 1 and row['PosicionTra'] == 1 and row['PosicionCab'] == 1:
				aux="Del Tra Cab"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 1 and row['PosicionCab'] == 0:
				aux="Del Tra"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 0 and row['PosicionCab'] == 1:
				aux="Del Cab"
			elif row['PosicionDel'] == 1 and row['PosicionTra'] == 0 and row['PosicionCab'] == 0:
				aux="Del"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 1 and row['PosicionCab'] == 1:
				aux="Tra Cab"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 1 and row['PosicionCab'] == 0:
				aux="Tra"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 0 and row['PosicionCab'] == 1:
				aux="Cab"
			elif row['PosicionDel'] == 0 and row['PosicionTra'] == 0 and row['PosicionCab'] == 0:
				aux="Ninguna"

			obj={
				'ID': row['ID'],
				'GABRIEL': row['GABRIEL'],
				'MONROE': row['MONROE'], 
				'GRC': row['GRC'], 
				'Armadora': row['Armadora'], 
				'Posicion': aux, 
				'Tipo': row['Tipo'], 
				'LongitudExp': row['LongitudExp'], 
				'LongitudComp': row['LongitudComp'], 
				'Carrera': row['Carrera'], 
				'TipoMontajeSup': row['TipoMontajeSup'], 
				'DiametroSup': row['DiametroSup'], 
				'LongitudSup': row['LongitudSup'], 
				'TipoMontajeInf': row['TipoMontajeInf'], 
				'DiametroInf': row['DiametroInf'], 
				'LongitudInf': row['LongitudInf'],
				"info": row['info']
			}
			result.append(obj)

		if result:
			#print(result)
			return {'intStatus':200, 'strAnswer': result}
		else:
			return {'intStatus':200, 'strAnswer': 'No hay nada'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnEditProductSP(GABRIEL,MONROE,GRC,Armadora,Posicion,Tipo,
                    LongitudExp,LongitudComp,Carrera,TipoMontajeSup,DiametroSup,LongitudSup,TipoMontajeInf,DiametroInf,LongitudInf,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()

		if Posicion == "DEL":
			PosicionDel = 1
			PosicionTra = 0
			PosicionCab = 0
		elif Posicion == "TRA":
			PosicionDel = 0
			PosicionTra = 1
			PosicionCab = 0
		elif Posicion == "CAB":
			PosicionDel = 0
			PosicionTra = 0
			PosicionCab = 1
		elif Posicion == "DEL TRA":
			PosicionDel = 1
			PosicionTra = 1
			PosicionCab = 0
		
		params = (
			GABRIEL,
			MONROE,
			GRC,
			Armadora,
			PosicionDel,
			PosicionTra,
			PosicionCab,
			Tipo,
			LongitudExp,
			LongitudComp,
			Carrera,
			TipoMontajeSup,
			DiametroSup,
			LongitudSup,
			TipoMontajeInf,
			DiametroInf,
			LongitudInf,
			info
		)
		cursor.callproc('putEditProductSP',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnDeleteProductSP(GABRIEL):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (GABRIEL,)
		cursor.callproc('deleteProductSP',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha eliminado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetBolsasDeAireID(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('getBolsasDeAireID',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnAddProductBA(GABRIEL,FIRESTONE,GOODYEAR,CONTITECH,Aplicacion1,Aplicacion2,Aplicacion3,OE1,OE2,OE3,Tapa,Membrana,Piston,inf):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			GABRIEL,
			FIRESTONE,
			GOODYEAR,
			CONTITECH,
			Aplicacion1,
			Aplicacion2,
			Aplicacion3,
			OE1,
			OE2,
			OE3,
			Tapa,
			Membrana,
			Piston,
			inf
		)
		cursor.callproc('postAddProductBA',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnEditProductBA(GABRIEL,FIRESTONE,GOODYEAR,CONTITECH,Aplicacion1,Aplicacion2,Aplicacion3,OE1,OE2,OE3,Tapa,Membrana,Piston,inf):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			GABRIEL,
			FIRESTONE,
			GOODYEAR,
			CONTITECH,
			Aplicacion1,
			Aplicacion2,
			Aplicacion3,
			OE1,
			OE2,
			OE3,
			Tapa,
			Membrana,
			Piston,
			inf
		)
		cursor.callproc('putEditProductBA',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnDeleteProductBA(GABRIEL):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (GABRIEL,)
		cursor.callproc('deleteProductBA',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha eliminado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetServicioLigeroID(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('getServicioLigeroID',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnAddProductSL(MarcaAuto,Submarca,Referencia,Modelo,AnoInicio,AnoFinal,Marca,Posicion,Tipo,LongExp,LongComp,Carrera,MontSup,MontInf,MONROE,GRC,KYB,BOGE,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			MarcaAuto,
			Submarca,
			Referencia,
			Modelo,
			AnoInicio,
			AnoFinal,
			Marca,
			Posicion,
			Tipo,
			LongExp,
			LongComp,
			Carrera,
			MontSup,
			MontInf,
			MONROE,
			GRC,
			KYB,
			BOGE,
			info
		)
		cursor.callproc('postAddProductSL',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnEditProductSL(MarcaAuto,Submarca,Referencia,Modelo,AnoInicio,AnoFinal,Marca,Posicion,Tipo,LongExp,LongComp,Carrera,
                    MontSup,MontInf,MONROE,GRC,KYB,BOGE,info,ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			ID,
			MarcaAuto,
			Submarca,
			Referencia,
			Modelo,
			AnoInicio,
			AnoFinal,
			Marca,
			Posicion,
			Tipo,
			LongExp,
			LongComp,
			Carrera,
			MontSup,
			MontInf,
			MONROE,
			GRC,
			KYB,
			BOGE,
			info
		)
		cursor.callproc('putEditProductSL',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnDeleteProductSL(MarcaAuto,Submarca,Modelo,Referencia):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (MarcaAuto,Submarca,Referencia,Modelo)
		cursor.callproc('deleteProductSL',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha eliminado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetMuellesID(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('getMuellesID',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnAddProductMue(filename):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		
		datosExcel = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Hoja1'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A','B','C','D','E','F','G','H','I','J','K','L']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
			datosExcel.append(datosfila)
		
		for row in datosExcel:
			params = (
				row[0], #RASSINI
				row[1], #MAF
				row[2], #SANDOVAL
				row[3], #ORIGINAL
				row[4], #No
				row[5], #Ancho
				row[6], #Espesor
				row[7], #Lfrontal
				row[8], #Ltrasero
				row[9], #Posicion
				row[10], #info
				row[11] #marca
			)
			cursor.callproc('postAddProductMue',params)

		MysqlCnx.commit()
		archivo_excel.close()
		os.remove(filename)
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnEditProductMue(RASSINI,MAF,SANDOVAL,ORIGINAL,No,Ancho,Espesor,Lfrontal,Ltrasero,Posicion,info,marca):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			RASSINI,
			MAF,
			SANDOVAL,
			ORIGINAL,
			No,
			Ancho,
			Espesor,
			Lfrontal,
			Ltrasero,
			Posicion,
			info,
			marca
		)
		cursor.callproc('putEditProductMue',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnDeleteProductMue(RASSINI):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (RASSINI,)
		cursor.callproc('deleteProductMue',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha eliminado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetTiposRefa():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		cursor.callproc('getRefaccionesTipos')
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetRefacciones(tipo):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()

		if tipo == "all":
			cursor.callproc('getRefaccionesAll')
		else:
			params = (
				tipo,
			)
			cursor.callproc('getRefacciones',params)

		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostRefacciones(ID,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			ID,
			info
		)
		cursor.callproc('postRefaccionesInfo',params)
		MysqlCnx.commit()

		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetRefaccionesID(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('getRefaccionesID',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnAddProductRefa(filename):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		
		datosExcel = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Hoja1'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
			datosExcel.append(datosfila)

		for datos in datosExcel:
			if datos[0] != None:
				params = (
					datos[0], #idModelo
					datos[1], #Descripcion
					datos[2], #TipoRefaccion
					datos[12], #TipoForma
					datos[3], #Unidad
					datos[4], #Modelo
					datos[5], #Anio
					datos[6], #Posicion
					datos[7], #DiametroInt
					datos[8], #DiametroExt
					datos[9], #Largo
					datos[10], #LargoTot
					datos[11] #info
				)
				cursor.callproc('postAddProductRefa',params)

		MysqlCnx.commit()
		archivo_excel.close()
		os.remove(filename)
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnEditProductRefa(ID,idModelo,Descripcion,Tipo,TipoForma,Unidad,Modelo,Anio,Posicion,DiametroInt,DiametroExt,Largo,LargoTot,info):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			ID,
			idModelo,
			Descripcion,
			Tipo,
			TipoForma,
			Unidad,
			Modelo,
			Anio,
			Posicion,
			DiametroInt,
			DiametroExt,
			Largo,
			LargoTot,
			info
		)
		cursor.callproc('putEditProductRefa',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha guardado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnDeleteProductRefa(ID):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (ID,)
		cursor.callproc('deleteProductRefa',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': "Se ha eliminado la informacion correctamente."}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetTornilleria(idCovamsa):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()

		params = (
			idCovamsa,
		)
		cursor.callproc('getTornilleria',params)
		response = cursor.fetchall()

		if response:
			params = (
				response[0]['Linea'],
			)
			cursor.callproc('getTabuladorPrecios',params)
			response2 = cursor.fetchall()

			costos=[float(resp['Costo']) for resp in response]
			costoMaximo = max(costos)
			costoMinimo = min(costos)
			proveedor_maximo = next(dato["Proveedor"] for dato in response if float(dato["Costo"]) == costoMaximo)
			proveedor_minimo = next(dato["Proveedor"] for dato in response if float(dato["Costo"]) == costoMinimo)
			answer = {
				"proveedores": response,
				"proveedor_maximo": proveedor_maximo,
				"costoMaximo": costoMaximo,
				"proveedor_minimo": proveedor_minimo,
				"costoMinimo": costoMinimo,
				"Publico": response2[0]['Publico'],
				"Mayoreo": response2[0]['Mayoreo'],
				"Platino": response2[0]['Platino'],
				"Comercio": response2[0]['Comercio'],
			}
			return {'intStatus':200, 'strAnswer': answer}	
		else:
			return {'intStatus':202, 'strAnswer': 'No hay proveedores.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetProvedores(tipoProv):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)
		cursor = MysqlCnx.cursor()
		params = (
			tipoProv,
		)
		cursor.callproc('getProveedores',params)
		response = cursor.fetchall()

		if response:
			return {'intStatus':200, 'strAnswer': response}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay proveedores.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetProductoProv(tabla,clave):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			tabla,
			clave
		)
		cursor.callproc('getProductoProv',params)
		response = cursor.fetchall()
		#print(response)

		if response:
			return {'intStatus':200, 'strAnswer': response}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay productos.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostNuevoCatalogo(filename,table):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()

		params=(table,)
		cursor.callproc('deleteCatalogoPrecios',params)

		datosExcel = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Hoja1'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A','B','C','D','E','F']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
			datosExcel.append(datosfila)

		fecha = datetime.datetime.now().strftime("%d-%m-%y")
		
		for fila in datosExcel:
			params = (
				table,
				fila[0],
				fila[1],
				fila[2],
				fila[3],
				fila[4],
				fila[5],
				str(fecha)
			)
			cursor.callproc('postNuevoCatalogoPrecios',params)

		MysqlCnx.commit()
		archivo_excel.close()
		os.remove(filename)
		return {'intStatus':200, 'strAnswer': 'Se ha guardado la informacion correctamente.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetTornilleriaProd(idCovamsa, proveedor, tabla):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			tabla,
			idCovamsa,
			proveedor
		)
		cursor.callproc('getTornilleriaProd',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay productos.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetLineasTor():
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		cursor.callproc('getLineasTor')
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay lineas.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetTabPrecios(linea):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			linea,
		)
		cursor.callproc('getTabuladorPrecios',params)
		response = cursor.fetchall()
		if response:
			return {'intStatus':200, 'strAnswer': response}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay productos.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetTornilleriaCostos(idCovamsa, linea, arrayProveedores):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		response = []
		params = (
			linea,
			idCovamsa
		)
		cursor.callproc('getTornilleriaLinea',params)
		ids = cursor.fetchall()

		for id in ids:
			aux = {
				"idCovamsa": id['idCovamsa'],
				"Descripcion": id['Descripcion'],
			}
			for proveedor in arrayProveedores:
				params = (
					id['idCovamsa'],
					proveedor
				)
				#print(params)
				cursor.callproc('getCostosTornilleria',params)
				costo = cursor.fetchall()
				if costo:
					aux.update({proveedor: float(costo[0]['Costo'])})
			response.append(aux)
		return {'intStatus':200, 'strAnswer': response}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()
	
def fnPostVerificarCoincidencias(filename, proveedor):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()

		datosExcel = []
		data = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Plantilla'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			if fila[hoja_trabajo[f"{columnasDeseadas[0]}1"].column - 1] != None:
				datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
				datosExcel.append(datosfila)

		for fila in datosExcel:
			params = (
				fila[0],
				proveedor
			)
			cursor.callproc('postVerificarCoincidencias',params)
			response = cursor.fetchall()
			if(response[0]['idCovamsa'] == None):
				aux = {
					"CLAVE": fila[0],
					"Nombre": response[0]['Nombre'],
				}
				data.append(aux)
		archivo_excel.close()
		os.remove(filename)
		return {'intStatus':200, 'strAnswer': data}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostAddCoincidencias(filename,proveedor):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()

		datosExcel = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Plantilla'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A','C']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			if fila[hoja_trabajo[f"{columnasDeseadas[0]}1"].column - 1] != None:
				datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
				datosExcel.append(datosfila)

		for fila in datosExcel:
			params = (
				fila[0],
				proveedor,
				fila[1]
			)
			cursor.callproc('postAddCoincidencias',params)

		MysqlCnx.commit()
		archivo_excel.close()
		os.remove(filename)

		return {'intStatus':200, 'strAnswer': 'Se ha guardado la informacion correctamente.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostAddProductoTor(filename,linea,proveedor):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()

		datosExcel = []
		archivo_excel = openpyxl.load_workbook(filename)
		nombreHoja = 'Plantilla'
		hoja_trabajo = archivo_excel[nombreHoja]

		columnasDeseadas = ['A','B','C']

		for fila in archivo_excel[nombreHoja].iter_rows(min_row=2,values_only=True):
			if fila[hoja_trabajo[f"{columnasDeseadas[0]}1"].column - 1] != None:
				datosfila = [fila[hoja_trabajo[f"{col}1"].column - 1] for col in columnasDeseadas]
				datosExcel.append(datosfila)

		for fila in datosExcel:
			params = (
				fila[0], #idCovamsa
				fila[1], #Descripcion
				linea, #Linea
				fila[2], #Clave
				proveedor	 #Proveedor
			)
			cursor.callproc('postAddProductTor',params)

		MysqlCnx.commit()
		archivo_excel.close()
		os.remove(filename)
		return {'intStatus':200, 'strAnswer': 'Se ha guardado la informacion correctamente.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetProdHerramientasProv(tabla,descripcion):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			tabla,
			descripcion
		)
		answer = []
		cursor.callproc('getProdHerramientasProv',params)
		response = cursor.fetchall()

		for resp in response:
			params = (
				resp['CLAVE'],
			)
			cursor.callproc('comprobarHerramienta',params)
			response2 = cursor.fetchall()
			if response2[0]['clave_existente'] != 1:
				aux = {
					"CLAVE": resp['CLAVE'],
					"Descripcion": resp['Descripcion']
				}
				answer.append(aux)

		if response:
			return {'intStatus':200, 'strAnswer': answer}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay productos.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPostHerramientasRelacion(descripcion,datos):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			descripcion,
		)
		cursor.callproc('postAddHerramientas',params)
		id = cursor.fetchall()
		for dato in datos:
			params = (
				id[0]['id'],
				dato['CLAVE'],
				dato['proveedor']
			)
			cursor.callproc('postAddRelacionHerra',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': 'Se ha guardado la informacion correctamente.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnGetHerramientasCoincidencia(CLAVE):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		params = (
			CLAVE,
		)
		cursor.callproc('getHerramientaID',params)
		id = cursor.fetchall()

		if id:
			params = (
				id[0]['idCovamsa'],
			)
			cursor.callproc('getHerramientasCoincidencia',params)
			response = cursor.fetchall()
			if response:
				return {'intStatus':200, 'strAnswer': response}
			else:
				return {'intStatus':202, 'strAnswer': 'No hay productos.'}
		else:
			return {'intStatus':202, 'strAnswer': 'No hay productos.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()

def fnPutHerramientasCoincidencia(idCovamsa,Proveedor,CLAVE):
	try:
		MysqlCnx = pymysql.connect(host=strMysqlHost,
						port=strMysqlPort,
						user=strMysqluUser,
						password=strMysqlPassword,
						db=strMysqlDB,
						charset='utf8mb4',
						cursorclass=pymysql.cursors.DictCursor)

		cursor = MysqlCnx.cursor()
		
		params = (
			idCovamsa,
			CLAVE,
			Proveedor
		)
		cursor.callproc('putHerramientasCoincidencia',params)
		MysqlCnx.commit()
		return {'intStatus':200, 'strAnswer': 'Se ha guardado la informacion correctamente.'}
	except Exception as e:
		return {'intStatus':500, 'strAnswer': str(e)}
	finally:
		MysqlCnx.close()